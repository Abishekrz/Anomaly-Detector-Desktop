# detection_core.py

import os
import datetime
from pathlib import Path

from streamlit import json
from inference.detector import load_models
from inference.commenter import generate_comments
from utils.viz import draw_boxes
from openpyxl import Workbook, load_workbook

# Base directory (Desktop_App/)
BASE_DIR = Path(__file__).parent

# Load YOLO models ONCE globally (Ultralytics YOLO v11 assumed)
cfg, models = load_models()


# -------------------------------------------------------------
#  Create new session folder: sessions/YYYY-MM-DD_HH-MM-SS/
# -------------------------------------------------------------
def create_session_folder():
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    session = BASE_DIR / "sessions" / now

    # Create subfolders
    (session / "uploads").mkdir(parents=True, exist_ok=True)
    (session / "results").mkdir(parents=True, exist_ok=True)

    print("\n===============================")
    print(" NEW SESSION:", session)
    print("===============================\n")

    return session


# -------------------------------------------------------------
# Save Excel results inside the session folder
# -------------------------------------------------------------
def save_to_excel(session_results_dir: Path, actual_image_path: str, annotated_image_path: str,
                  detections: list, comments: list):
    """
    Save one row per inference to results.xlsx inside session_results_dir.

    Columns:
    - Timestamp
    - Actual Image Path
    - Actual Image Name
    - Annotated Image Path
    - Annotated Image Name
    - Findings (JSON string of detections)
    - Comments (semi-colon separated)
    """
    excel_path = session_results_dir / "results.xlsx"
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Prepare data
    actual_image_path = str(actual_image_path)
    annotated_image_path = str(annotated_image_path)
    actual_image_name = Path(actual_image_path).name
    annotated_image_name = Path(annotated_image_path).name

    # Serialize detections to JSON (compact)
    try:
        findings_json = json.dumps(detections, ensure_ascii=False)
    except Exception:
        # Fallback: simple string representation
        findings_json = str(detections)

    comments_text = "; ".join(comments) if comments else ""

    # Create or append workbook
    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append([
            "Timestamp",
            "Actual Image Path",
            "Actual Image Name",
            "Annotated Image Path",
            "Annotated Image Name",
            "Findings (JSON)",
            "Comments"
        ])

    ws.append([now, actual_image_path, actual_image_name, annotated_image_path,
               annotated_image_name, findings_json, comments_text])
    wb.save(excel_path)

    print("EXCEL SAVED:", excel_path)


# -------------------------------------------------------------
# Main function: run YOLO models on image and produce results
# Accepts optional enabled_models dict (name -> model_object)
# If enabled_models is None, uses the global 'models'
# -------------------------------------------------------------
def run_inference_on_path(image_path: str, session_results_dir: Path, enabled_models: dict = None):
    image_path = str(image_path).replace("\\", "/")
    print("\nINPUT IMAGE:", image_path)

    # Choose which models to run
    run_models = enabled_models if (enabled_models is not None and len(enabled_models) > 0) else models

    all_detections = []

    # ------------ Run each available model ------------- #
    for model_name, model_obj in run_models.items():
        print(f"\nRunning model '{model_name}' on image...")

        # Execute prediction (Ultralytics v11 -> Results objects)
        try:
            results = model_obj.predict(image_path)  # returns Results object(s)
        except Exception as e:
            print(f"ERROR running model '{model_name}':", e)
            results = []

        parsed_dets = []

        # Parse Ultralytics Results -> list of detection dicts
        for r in results:
            # r.boxes might be None or empty
            boxes = getattr(r, "boxes", None)
            names = getattr(r, "names", {})  # mapping id->name
            if boxes is None:
                continue

            # Each box in r.boxes is a Box object with attributes xyxy, conf, cls
            for b in boxes:
                try:
                    # xyxy, conf, cls are tensors, access [0] then convert
                    xyxy_tensor = b.xyxy[0]  # tensor-like
                    xyxy = [int(round(float(x))) for x in xyxy_tensor.tolist()]

                    conf_val = float(b.conf[0]) if getattr(b, "conf", None) is not None else 0.0
                    cls_id = int(b.cls[0]) if getattr(b, "cls", None) is not None else -1

                    label = names.get(cls_id, str(cls_id)) if isinstance(names, dict) else str(cls_id)

                    parsed_dets.append({
                        "bbox": xyxy,           # [x1, y1, x2, y2] ints
                        "confidence": conf_val, # float
                        "label": label,         # string
                        "model": model_name
                    })
                except Exception as e:
                    print("Box parsing error (skipping one box):", e)
                    continue

        print(f"Model '{model_name}' detections:", len(parsed_dets))
        all_detections.extend(parsed_dets)

    # ------------ Generate comments ------------ #
    try:
        comments = generate_comments(all_detections)
    except Exception as e:
        print("COMMENT GENERATION ERROR:", e)
        comments = ["Error generating comments"]

    # ------------ Save annotated image ------------ #
    out_name = f"annotated_{Path(image_path).name}"
    out_path = session_results_dir / out_name
    out_path_str = str(out_path).replace("\\", "/")

    print("OUTPUT FILE PATH:", out_path_str)

    try:
        draw_boxes(image_path, all_detections, out_path_str)
        print("ANNOTATION SAVED:", out_path_str)
    except Exception as e:
        print("DRAW ERROR:", e)

    # ------------ Save Excel ------------ #
        # ------------ Save Excel (NEW call signature) ------------ #
    try:
        save_to_excel(session_results_dir, image_path, out_path_str, all_detections, comments)
    except Exception as e:
        print("EXCEL SAVE ERROR:", e)


    return out_path_str, all_detections, comments


# -------------------------------------------------------------
# Basic folder structure creation
# -------------------------------------------------------------
def ensure_dirs():
    (BASE_DIR / "sessions").mkdir(exist_ok=True)
